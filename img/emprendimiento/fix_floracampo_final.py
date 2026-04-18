"""
FLORACAMPO - Corrección Completa del Plan Financiero
Todos los valores son ANUALES (no mensuales)
Estándar: Fondo Emprender / SENA 2025
"""

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
import copy

FILE = 'PLAN DE NEGOCIOS FORMULAS.xlsx'

# ─── Paleta de colores corporativa FLORACAMPO ──────────────────────────────
VERDE_OSCURO  = "1B5E20"   # encabezados principales
VERDE_MEDIO   = "2E7D32"   # sub‑encabezados
VERDE_CLARO   = "A5D6A7"   # filas de categoría
AMARILLO      = "FFF9C4"   # filas de detalle
BLANCO        = "FFFFFF"
GRIS_CLARO    = "F5F5F5"
NARANJA       = "E65100"   # totales / alertas
NARANJA_CLARO = "FFE0B2"   # fondo totales

# ─── Helpers ───────────────────────────────────────────────────────────────
def borde():
    s = Side(style='thin', color='BDBDBD')
    return Border(left=s, right=s, top=s, bottom=s)

def borde_grueso():
    s = Side(style='medium', color='1B5E20')
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def header_cell(ws, row, col, value, color=VERDE_OSCURO, font_color="FFFFFF",
                bold=True, size=11, border=True, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(color)
    c.font = Font(bold=bold, color=font_color, size=size, name="Calibri")
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if border:
        c.border = borde()
    return c

def detail_cell(ws, row, col, value, bg=BLANCO, bold=False, align="left",
                number_fmt=None, font_color="212121"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(bg)
    c.font = Font(bold=bold, color=font_color, size=10, name="Calibri")
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = borde()
    if number_fmt:
        c.number_format = number_fmt
    return c

PESO = '#,##0'       # formato moneda colombiana sin decimales
PCT  = '0.0%'
NUM  = '#,##0'

YEARS = [2025, 2026, 2027, 2028, 2029]
COLS_AÑO = ['B', 'C', 'D', 'E', 'F']   # columna por año

# ═══════════════════════════════════════════════════════════════════════════
# DATOS BASE DEL MODELO
# ═══════════════════════════════════════════════════════════════════════════
# INGRESOS anuales
# Frutas Amazónicas: 6.000 kg/mes × 12 = 72.000 kg/año | precio $9.500/kg
# Hortalizas:        4.000 kg/mes × 12 = 48.000 kg/año | precio $4.500/kg
# Crec. unidades 8 %/año · crec. precio 5 %/año (IPC estimado)

P1_U  = 72_000    # kg año 1
P1_P  = 9_500     # COP/kg año 1
P2_U  = 48_000
P2_P  = 4_500
CREC_U = 0.08
CREC_P = 0.05

def proj_ingresos():
    """Devuelve lista de dicts con proyecciones por año."""
    rows = []
    for j in range(5):
        u1  = int(P1_U  * (1 + CREC_U) ** j)
        pr1 = int(P1_P  * (1 + CREC_P) ** j)
        u2  = int(P2_U  * (1 + CREC_U) ** j)
        pr2 = int(P2_P  * (1 + CREC_P) ** j)
        rows.append({
            'año': YEARS[j],
            'u1': u1, 'pr1': pr1, 'tot1': u1 * pr1,
            'u2': u2, 'pr2': pr2, 'tot2': u2 * pr2,
            'total': u1 * pr1 + u2 * pr2,
        })
    return rows

# NOMINA (anual) — SMMLV 2025 = $1,423,500
# Cargo                    N   Factor prestaciones  Total/mes/persona    → Total/año
# Director operativo       1   1.60                 2,277,600            27,331,200
# Coordinador bodega       1   1.50                 2,135,250            25,623,000
# Operarios clasificación  3   1.40                 1,992,900 c/u        23,914,800 ×3 = 71,744,400
# Conductores/logística    2   1.45                 2,064,075 c/u        24,768,900 ×2 = 49,537,800
# Asesor comercial         1   1.50                 2,135,250            25,623,000
# Gestor comunitario       1   1.40                 1,992,900            23,914,800
# Personal TI              1   1.55                 2,206,425            26,477,100
# Técnico mantenimiento    1   1.40                 1,992,900            23,914,800
# TOTAL año 1 (10 personas):                                            274,166,100

NOMINA_Y1 = 274_166_100
CREC_NOM   = 0.06          # incremento salarial ≈ inflación proyectada

def proj_nomina():
    return [int(NOMINA_Y1 * (1 + CREC_NOM) ** j) for j in range(5)]

# EGRESOS OPERATIVOS (anuales)
# Materia prima / insumos: 30 % de ventas
# Combustible y fletes:    $10M/mes × 12 = $120M año 1, crece 6 %
# Energía y mantenimiento: $5M/mes × 12  = $60M  año 1, crece 6 %
# Seguros                : $2.5M/mes× 12 = $30M  año 1, crece 5 %
# Publicidad             : $1.5M/mes× 12 = $18M  año 1, crece 5 %
# Empaque sostenible     : $0.8M/mes× 12 = $9.6M año 1, crece 5 %
# Distribucion                            $12M   año 1, crece 5 %

def proj_egresos(ing):
    rows = []
    for j in range(5):
        ventas = ing[j]['total']
        rows.append({
            'año': YEARS[j],
            'nomina'      : proj_nomina()[j],
            'mat_prima'   : int(ventas * 0.28),
            'combustible' : int(120_000_000 * (1.06 ** j)),
            'energia'     : int(60_000_000  * (1.06 ** j)),
            'seguros'     : int(30_000_000  * (1.05 ** j)),
            'publicidad'  : int(18_000_000  * (1.05 ** j)),
            'empaque'     : int(9_600_000   * (1.05 ** j)),
            'distribucion': int(12_000_000  * (1.05 ** j)),
        })
    return rows

# INVERSIÓN INICIAL (solo año 1, reinversión de mantenimiento años 2‑5)
INV = {
    'Infraestructura (bodega + adecuaciones)'  : 120_000_000,
    'Equipos de frío y empaque'                : 150_000_000,
    'Camiones de carga (2 unidades)'           : 130_000_000,
    'Software FLOR-Net (plataforma digital)'   : 30_000_000,
    'Equipos de cómputo y redes'               : 16_000_000,
    'Muebles y enseres'                        : 12_000_000,
    'Constitución legal y permisos INVIMA'     : 8_500_000,
    'Capital de trabajo inicial (3 meses)'     : 68_000_000,
    'Caja menor y fondo de emergencia'         : 15_000_000,
}
# Mantenimiento anual 5 % de activos físicos (años 2‑5)
ACTIVOS_FISICOS = (120_000_000 + 150_000_000 + 130_000_000 +
                   16_000_000 + 12_000_000)
MANT_ANUAL = int(ACTIVOS_FISICOS * 0.05)

# PLANTA DE PERSONAL (número de personas)
PERSONAL = {
    'Área Administrativa'         : [2, 2, 3, 3, 4],
    'Mercadeo y Ventas'           : [1, 2, 2, 3, 3],
    'Producción y Operaciones'    : [5, 6, 7, 8, 9],
    'Áreas Transversales (TI/Mant)': [2, 2, 2, 3, 3],
}

# COSTOS FABRICACIÓN MENSUAL (para la hoja de referencia)
COS_FAB_M = {
    'Nómina'       : 22_847_175,   # NOMINA_Y1 / 12
    'Distribución' : 10_000_000,
    'Materia Prima': None,         # variable → % ventas
    'Mano de Obra Indirecta': 3_500_000,
    'Publicidad'   : 1_500_000,
    'Servicios'    : 5_000_000,
    'Etiquetas/Empaque': 800_000,
}

# ═══════════════════════════════════════════════════════════════════════════
# ESCRITURA DEL WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════

def write_proyeccion_ventas(ws, ing):
    ws.title = "PROYECCION DE VENTAS A 5 AÑOS"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 42
    for c in COLS_AÑO:
        ws.column_dimensions[c].width = 18

    # Título
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = "PROYECCIÓN DE VENTAS A 5 AÑOS — FLORACAMPO"
    c.fill = _fill(VERDE_OSCURO); c.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Sección precios
    ws.merge_cells('A2:F2')
    c = ws['A2']; c.value = "1. PROYECCIÓN DE PRECIOS POR PRODUCTO (COP/kg)"
    c.fill = _fill(VERDE_MEDIO); c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center"); ws.row_dimensions[2].height = 20

    header_cell(ws, 3, 1, "PRODUCTO", VERDE_OSCURO)
    for j, yr in enumerate(YEARS):
        header_cell(ws, 3, j + 2, str(yr), VERDE_OSCURO)

    detail_cell(ws, 4, 1, "Frutas Amazónicas (Copoazú, Arazá, Sacha Inchi)", AMARILLO, bold=True)
    detail_cell(ws, 5, 1, "Hortalizas (Picado/Empaque Segunda Transformación)", AMARILLO, bold=True)
    detail_cell(ws, 6, 1, "Precio promedio ponderado (COP/kg)", VERDE_CLARO, bold=True)

    for j in range(5):
        col = j + 2
        detail_cell(ws, 4, col, ing[j]['pr1'], AMARILLO, align="right", number_fmt=PESO)
        detail_cell(ws, 5, col, ing[j]['pr2'], AMARILLO, align="right", number_fmt=PESO)
        pp = (ing[j]['tot1'] + ing[j]['tot2']) / (ing[j]['u1'] + ing[j]['u2'])
        detail_cell(ws, 6, col, int(pp), VERDE_CLARO, bold=True, align="right", number_fmt=PESO)

    # Sección cantidades y ventas
    ws.merge_cells('A8:F8')
    c = ws['A8']; c.value = "2. PROYECCIÓN DE VENTAS ANUALES (UNIDADES × PRECIO = INGRESOS)"
    c.fill = _fill(VERDE_MEDIO); c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center"); ws.row_dimensions[8].height = 20

    header_cell(ws, 9, 1, "CONCEPTO")
    for j, yr in enumerate(YEARS):
        header_cell(ws, 9, j + 2, str(yr))

    labels = [
        ("Frutas Amazónicas — Unidades (kg/año)", None),
        ("Frutas Amazónicas — Precio (COP/kg)",   None),
        ("Frutas Amazónicas — TOTAL INGRESOS",    "tot1"),
        ("",                                       None),
        ("Hortalizas — Unidades (kg/año)",         None),
        ("Hortalizas — Precio (COP/kg)",           None),
        ("Hortalizas — TOTAL INGRESOS",            "tot2"),
        ("",                                       None),
        ("★  TOTAL VENTAS ANUALES",               "total"),
    ]
    for i, (label, key) in enumerate(labels):
        r = 10 + i
        bg = NARANJA_CLARO if key == "total" else (VERDE_CLARO if key in ("tot1","tot2") else GRIS_CLARO if i % 2 == 0 else BLANCO)
        bold = key in ("total", "tot1", "tot2")
        fc = NARANJA if key == "total" else "212121"
        detail_cell(ws, r, 1, label, bg, bold=bold, font_color=fc)
        if not label:
            continue
        for j in range(5):
            col = j + 2
            if key == "total":
                v = ing[j]['total']
                detail_cell(ws, r, col, v, NARANJA_CLARO, bold=True, align="right", number_fmt=PESO, font_color=NARANJA)
            elif key == "tot1":
                detail_cell(ws, r, col, ing[j]['tot1'], VERDE_CLARO, bold=True, align="right", number_fmt=PESO)
            elif key == "tot2":
                detail_cell(ws, r, col, ing[j]['tot2'], VERDE_CLARO, bold=True, align="right", number_fmt=PESO)
            elif "Unidades" in label and "Frutas" in label:
                detail_cell(ws, r, col, ing[j]['u1'], bg, align="right", number_fmt=NUM)
            elif "Precio" in label and "Frutas" in label:
                detail_cell(ws, r, col, ing[j]['pr1'], bg, align="right", number_fmt=PESO)
            elif "Unidades" in label and "Horta" in label:
                detail_cell(ws, r, col, ing[j]['u2'], bg, align="right", number_fmt=NUM)
            elif "Precio" in label and "Horta" in label:
                detail_cell(ws, r, col, ing[j]['pr2'], bg, align="right", number_fmt=PESO)


def write_costos_fabricacion(ws):
    ws.title = "COSTOS DE FABRICACION MENSUAL"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 20

    ws.merge_cells('A1:B1')
    c = ws['A1']; c.value = "COSTOS DE FABRICACIÓN MENSUAL — REFERENCIA AÑO 1"
    c.fill = _fill(VERDE_OSCURO); c.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 30

    header_cell(ws, 2, 1, "CONCEPTO"); header_cell(ws, 2, 2, "VALOR MENSUAL (COP)")

    rows_fab = [
        ("COSTOS INDIRECTOS O FIJOS", None, VERDE_CLARO, True),
        ("Nómina total (10 personas)", int(NOMINA_Y1 / 12), GRIS_CLARO, False),
        ("Distribución y logística base", 10_000_000, GRIS_CLARO, False),
        ("TOTAL COSTOS FIJOS", int(NOMINA_Y1 / 12) + 10_000_000, VERDE_CLARO, True),
        ("", None, BLANCO, False),
        ("COSTOS DIRECTOS O VARIABLES", None, VERDE_CLARO, True),
        ("Materia prima (estimado promedio mensual)", int(28 * 7_750_000 / 100), GRIS_CLARO, False),
        ("Mano de obra indirecta / jornales", 3_500_000, GRIS_CLARO, False),
        ("Publicidad y marketing digital", 1_500_000, GRIS_CLARO, False),
        ("Servicios públicos (energía, agua, telecom)", 5_000_000, GRIS_CLARO, False),
        ("Etiquetas, empaques y materiales", 800_000, GRIS_CLARO, False),
        ("TOTAL COSTOS VARIABLES", int(28*7_750_000/100)+3_500_000+1_500_000+5_000_000+800_000, VERDE_CLARO, True),
        ("", None, BLANCO, False),
        ("★  TOTAL COSTOS MENSUALES", int(NOMINA_Y1/12)+10_000_000+int(28*7_750_000/100)+3_500_000+1_500_000+5_000_000+800_000, NARANJA_CLARO, True),
    ]

    for i, (label, value, bg, bold) in enumerate(rows_fab):
        r = 3 + i
        fc = NARANJA if "★" in label else "212121"
        detail_cell(ws, r, 1, label, bg, bold=bold, font_color=fc)
        if value is not None:
            detail_cell(ws, r, 2, value, bg, bold=bold, align="right", number_fmt=PESO, font_color=fc)


def write_planta_personal(ws):
    ws.title = "PLANTA DE PERSONAL"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 36
    for c in COLS_AÑO:
        ws.column_dimensions[c].width = 12

    ws.merge_cells('A1:F1')
    c = ws['A1']; c.value = "PLANTA DE PERSONAL — EVOLUCIÓN 2025–2029"
    c.fill = _fill(VERDE_OSCURO); c.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 30

    header_cell(ws, 2, 1, "ÁREA / CARGO")
    for j, yr in enumerate(YEARS):
        header_cell(ws, 2, j + 2, str(yr))

    detalle_personal = [
        ("ÁREA ADMINISTRATIVA",        [2, 2, 3, 3, 4], VERDE_CLARO, True),
        ("  Director Operativo",        [1, 1, 1, 1, 1], GRIS_CLARO, False),
        ("  Coordinador de Bodega",     [1, 1, 1, 1, 1], GRIS_CLARO, False),
        ("  Asistente Administrativo",  [0, 0, 1, 1, 2], GRIS_CLARO, False),
        ("MERCADEO Y VENTAS",           [1, 2, 2, 3, 3], VERDE_CLARO, True),
        ("  Asesor Comercial",          [1, 1, 2, 2, 2], GRIS_CLARO, False),
        ("  Gestor Comunitario",        [0, 1, 0, 1, 1], GRIS_CLARO, False),
        ("PRODUCCIÓN Y OPERACIONES",    [5, 6, 7, 8, 9], VERDE_CLARO, True),
        ("  Operarios Clasificación",   [3, 3, 4, 4, 5], GRIS_CLARO, False),
        ("  Conductores / Logística",   [2, 2, 2, 3, 3], GRIS_CLARO, False),
        ("  Técnico Agropecuario",      [0, 1, 1, 1, 1], GRIS_CLARO, False),
        ("ÁREAS TRANSVERSALES",         [2, 2, 2, 3, 3], VERDE_CLARO, True),
        ("  Desarrollador / TI",        [1, 1, 1, 1, 1], GRIS_CLARO, False),
        ("  Técnico Mantenimiento",     [1, 1, 1, 1, 1], GRIS_CLARO, False),
        ("  Coordinador de Calidad",    [0, 0, 0, 1, 1], GRIS_CLARO, False),
        ("★  TOTAL PERSONAL",           [10, 12, 14, 17, 19], NARANJA_CLARO, True),
    ]

    for i, (label, vals, bg, bold) in enumerate(detalle_personal):
        r = 3 + i
        fc = NARANJA if "★" in label else "212121"
        detail_cell(ws, r, 1, label, bg, bold=bold, font_color=fc)
        for j, v in enumerate(vals):
            detail_cell(ws, r, j + 2, v if v > 0 else "—", bg, bold=bold, align="center", font_color=fc)


def write_inversion_total(ws):
    ws.title = "INVERSION TOTAL"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 42
    for c in COLS_AÑO:
        ws.column_dimensions[c].width = 18

    ws.merge_cells('A1:F1')
    c = ws['A1']; c.value = "INVERSIÓN TOTAL — FLORACAMPO 2025–2029 (COP)"
    c.fill = _fill(VERDE_OSCURO); c.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 30

    header_cell(ws, 2, 1, "CONCEPTO")
    for j, yr in enumerate(YEARS):
        header_cell(ws, 2, j + 2, str(yr))

    ws.merge_cells('A3:F3')
    c = ws['A3']; c.value = "ACTIVOS FIJOS"
    c.fill = _fill(VERDE_MEDIO); c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")

    activos_fijos = [
        ("Infraestructura (bodega + adecuaciones)", 120_000_000, 0, 0, 0, 0),
        ("Equipos de frío y empaque", 150_000_000, 0, 0, 0, 0),
        ("Camiones de carga (2 unidades)", 130_000_000, 0, 0, 0, 65_000_000),   # renovación año 5
        ("Equipos de cómputo y redes", 16_000_000, 6_000_000, 0, 0, 0),
        ("Muebles y enseres", 12_000_000, 0, 0, 4_000_000, 0),
    ]
    subtot_fijos = [0]*5
    for i, row_data in enumerate(activos_fijos):
        r = 4 + i
        bg = GRIS_CLARO if i % 2 == 0 else BLANCO
        detail_cell(ws, r, 1, row_data[0], bg)
        for j in range(5):
            v = row_data[j+1]
            subtot_fijos[j] += v
            detail_cell(ws, r, j+2, v if v else "—", bg, align="right", number_fmt=PESO)

    r_sf = 4 + len(activos_fijos)
    detail_cell(ws, r_sf, 1, "Subtotal Activos Fijos", VERDE_CLARO, bold=True)
    for j in range(5):
        detail_cell(ws, r_sf, j+2, subtot_fijos[j], VERDE_CLARO, bold=True, align="right", number_fmt=PESO)

    ws.merge_cells(f'A{r_sf+1}:F{r_sf+1}')
    c = ws[f'A{r_sf+1}']; c.value = "ACTIVOS DIFERIDOS"
    c.fill = _fill(VERDE_MEDIO); c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")

    diferidos = [
        ("Software FLOR-Net (licencia y desarrollo)", 30_000_000, 5_000_000, 3_000_000, 3_000_000, 3_000_000),
        ("Constitución legal y permisos INVIMA",       8_500_000, 0, 0, 2_000_000, 0),
    ]
    subtot_dif = [0]*5
    for i, row_data in enumerate(diferidos):
        r = r_sf + 2 + i
        bg = GRIS_CLARO if i % 2 == 0 else BLANCO
        detail_cell(ws, r, 1, row_data[0], bg)
        for j in range(5):
            v = row_data[j+1]; subtot_dif[j] += v
            detail_cell(ws, r, j+2, v if v else "—", bg, align="right", number_fmt=PESO)

    r_sd = r_sf + 2 + len(diferidos)
    detail_cell(ws, r_sd, 1, "Subtotal Activos Diferidos", VERDE_CLARO, bold=True)
    for j in range(5):
        detail_cell(ws, r_sd, j+2, subtot_dif[j], VERDE_CLARO, bold=True, align="right", number_fmt=PESO)

    ws.merge_cells(f'A{r_sd+1}:F{r_sd+1}')
    c = ws[f'A{r_sd+1}']; c.value = "CAPITAL DE TRABAJO"
    c.fill = _fill(VERDE_MEDIO); c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")

    capital = [
        ("Capital de trabajo inicial (3 meses op.)", 68_000_000, 0, 0, 0, 0),
        ("Caja menor y fondo de emergencia",          15_000_000, 5_000_000, 5_000_000, 5_000_000, 5_000_000),
    ]
    subtot_ct = [0]*5
    for i, row_data in enumerate(capital):
        r = r_sd + 2 + i
        bg = GRIS_CLARO if i % 2 == 0 else BLANCO
        detail_cell(ws, r, 1, row_data[0], bg)
        for j in range(5):
            v = row_data[j+1]; subtot_ct[j] += v
            detail_cell(ws, r, j+2, v if v else "—", bg, align="right", number_fmt=PESO)

    r_sct = r_sd + 2 + len(capital)
    detail_cell(ws, r_sct, 1, "Subtotal Capital de Trabajo", VERDE_CLARO, bold=True)
    for j in range(5):
        detail_cell(ws, r_sct, j+2, subtot_ct[j], VERDE_CLARO, bold=True, align="right", number_fmt=PESO)

    r_inv = r_sct + 1
    ws.row_dimensions[r_inv].height = 22
    detail_cell(ws, r_inv, 1, "★  INVERSIÓN TOTAL", NARANJA_CLARO, bold=True, font_color=NARANJA)
    for j in range(5):
        tot = subtot_fijos[j] + subtot_dif[j] + subtot_ct[j]
        detail_cell(ws, r_inv, j+2, tot, NARANJA_CLARO, bold=True, align="right", number_fmt=PESO, font_color=NARANJA)


def write_flujo_egresos(ws, ing, egr):
    ws.title = "FLUJO DE EGRESOS"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 42
    for c in COLS_AÑO:
        ws.column_dimensions[c].width = 18

    ws.merge_cells('A1:F1')
    c = ws['A1']; c.value = "FLUJO DE EGRESOS ANUALES — FLORACAMPO 2025–2029 (COP)"
    c.fill = _fill(VERDE_OSCURO); c.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 30

    header_cell(ws, 2, 1, "CONCEPTO")
    for j, yr in enumerate(YEARS):
        header_cell(ws, 2, j+2, str(yr))

    # Sección Nómina
    ws.merge_cells('A3:F3')
    c = ws['A3']; c.value = "1. NÓMINA Y RECURSO HUMANO"
    c.fill = _fill(VERDE_MEDIO); c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")

    nom_rows = [
        ("Nómina total (incluyendo prestaciones y s. social)", "nomina"),
    ]
    for i, (label, key) in enumerate(nom_rows):
        r = 4 + i
        detail_cell(ws, r, 1, label, GRIS_CLARO)
        for j in range(5):
            detail_cell(ws, r, j+2, egr[j][key], GRIS_CLARO, align="right", number_fmt=PESO)

    r_nom = 5
    detail_cell(ws, r_nom, 1, "Subtotal Nómina", VERDE_CLARO, bold=True)
    for j in range(5):
        detail_cell(ws, r_nom, j+2, egr[j]['nomina'], VERDE_CLARO, bold=True, align="right", number_fmt=PESO)

    # Sección Costos Producción
    ws.merge_cells(f'A6:F6')
    c = ws['A6']; c.value = "2. COSTOS DE PRODUCCIÓN Y LOGÍSTICA"
    c.fill = _fill(VERDE_MEDIO); c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")

    prod_rows = [
        ("Materia prima e insumos (28 % de ventas)",   "mat_prima"),
        ("Combustible y fletes",                        "combustible"),
        ("Energía eléctrica y mantenimiento equipos",   "energia"),
        ("Seguros (carga, vehículos, infraestructura)", "seguros"),
    ]
    subtot_prod = [0]*5
    for i, (label, key) in enumerate(prod_rows):
        r = 7 + i
        bg = GRIS_CLARO if i % 2 == 0 else BLANCO
        detail_cell(ws, r, 1, label, bg)
        for j in range(5):
            v = egr[j][key]; subtot_prod[j] += v
            detail_cell(ws, r, j+2, v, bg, align="right", number_fmt=PESO)

    r_sp = 7 + len(prod_rows)
    detail_cell(ws, r_sp, 1, "Subtotal Producción y Logística", VERDE_CLARO, bold=True)
    for j in range(5):
        detail_cell(ws, r_sp, j+2, subtot_prod[j], VERDE_CLARO, bold=True, align="right", number_fmt=PESO)

    # Sección Comercialización
    ws.merge_cells(f'A{r_sp+1}:F{r_sp+1}')
    c = ws[f'A{r_sp+1}']; c.value = "3. COSTOS DE COMERCIALIZACIÓN Y VENTAS"
    c.fill = _fill(VERDE_MEDIO); c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")

    comm_rows = [
        ("Publicidad y marketing digital",  "publicidad"),
        ("Empaques sostenibles",            "empaque"),
        ("Distribución y entregas",         "distribucion"),
    ]
    subtot_comm = [0]*5
    for i, (label, key) in enumerate(comm_rows):
        r = r_sp + 2 + i
        bg = GRIS_CLARO if i % 2 == 0 else BLANCO
        detail_cell(ws, r, 1, label, bg)
        for j in range(5):
            v = egr[j][key]; subtot_comm[j] += v
            detail_cell(ws, r, j+2, v, bg, align="right", number_fmt=PESO)

    r_sc = r_sp + 2 + len(comm_rows)
    detail_cell(ws, r_sc, 1, "Subtotal Comercialización y Ventas", VERDE_CLARO, bold=True)
    for j in range(5):
        detail_cell(ws, r_sc, j+2, subtot_comm[j], VERDE_CLARO, bold=True, align="right", number_fmt=PESO)

    # TOTAL EGRESOS
    r_tot = r_sc + 1
    ws.row_dimensions[r_tot].height = 22
    detail_cell(ws, r_tot, 1, "★  TOTAL EGRESOS ANUALES", NARANJA_CLARO, bold=True, font_color=NARANJA)
    for j in range(5):
        tot = egr[j]['nomina'] + subtot_prod[j] + subtot_comm[j]
        detail_cell(ws, r_tot, j+2, tot, NARANJA_CLARO, bold=True, align="right", number_fmt=PESO, font_color=NARANJA)

    return subtot_prod, subtot_comm


def write_flujo_ingresos(ws, ing):
    ws.title = "FLUJO DE INGRESOS"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 46
    for c in COLS_AÑO:
        ws.column_dimensions[c].width = 18

    ws.merge_cells('A1:F1')
    c = ws['A1']; c.value = "FLUJO DE INGRESOS ANUALES — FLORACAMPO 2025–2029 (COP)"
    c.fill = _fill(VERDE_OSCURO); c.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 30

    header_cell(ws, 2, 1, "CONCEPTO")
    for j, yr in enumerate(YEARS):
        header_cell(ws, 2, j+2, str(yr))

    # P1 frutas
    ws.merge_cells('A3:F3')
    c = ws['A3']; c.value = "PRODUCTO 1 — Frutas Amazónicas (Copoazú, Arazá, Sacha Inchi)"
    c.fill = _fill(VERDE_MEDIO); c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")

    detail_cell(ws, 4, 1, "Unidades vendidas (kg/año)", GRIS_CLARO)
    detail_cell(ws, 5, 1, "Precio unitario (COP/kg)", BLANCO)
    detail_cell(ws, 6, 1, "TOTAL INGRESOS Producto 1", VERDE_CLARO, bold=True)

    for j in range(5):
        detail_cell(ws, 4, j+2, ing[j]['u1'], GRIS_CLARO, align="right", number_fmt=NUM)
        detail_cell(ws, 5, j+2, ing[j]['pr1'], BLANCO, align="right", number_fmt=PESO)
        detail_cell(ws, 6, j+2, ing[j]['tot1'], VERDE_CLARO, bold=True, align="right", number_fmt=PESO)

    # P2 hortalizas
    ws.merge_cells('A7:F7')
    c = ws['A7']; c.value = "PRODUCTO 2 — Hortalizas de Segunda Transformación (Picado/Empaque)"
    c.fill = _fill(VERDE_MEDIO); c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")

    detail_cell(ws, 8, 1, "Unidades vendidas (kg/año)", GRIS_CLARO)
    detail_cell(ws, 9, 1, "Precio unitario (COP/kg)", BLANCO)
    detail_cell(ws, 10, 1, "TOTAL INGRESOS Producto 2", VERDE_CLARO, bold=True)

    for j in range(5):
        detail_cell(ws, 8, j+2, ing[j]['u2'], GRIS_CLARO, align="right", number_fmt=NUM)
        detail_cell(ws, 9, j+2, ing[j]['pr2'], BLANCO, align="right", number_fmt=PESO)
        detail_cell(ws, 10, j+2, ing[j]['tot2'], VERDE_CLARO, bold=True, align="right", number_fmt=PESO)

    # Total
    ws.row_dimensions[11].height = 22
    detail_cell(ws, 11, 1, "★  TOTAL INGRESOS ANUALES", NARANJA_CLARO, bold=True, font_color=NARANJA)
    for j in range(5):
        detail_cell(ws, 11, j+2, ing[j]['total'], NARANJA_CLARO, bold=True, align="right", number_fmt=PESO, font_color=NARANJA)

    # Fuente de tracción
    ws.merge_cells('A12:F12')
    c = ws['A12']; c.value = "Supuestos: Crecimiento unidades 8 %/año | Incremento precio 5 %/año (IPC proyectado)"
    c.fill = _fill(GRIS_CLARO); c.font = Font(italic=True, color="555555", size=9, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")


def write_estado_resultados(ws, ing, egr, subtot_prod, subtot_comm):
    ws.title = "ESTADO DE RESULTADOS"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 44
    for c in COLS_AÑO:
        ws.column_dimensions[c].width = 18

    ws.merge_cells('A1:F1')
    c = ws['A1']; c.value = "ESTADO DE RESULTADOS PROYECTADO — FLORACAMPO 2025–2029 (COP)"
    c.fill = _fill(VERDE_OSCURO); c.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 30

    header_cell(ws, 2, 1, "CONCEPTO")
    for j, yr in enumerate(YEARS):
        header_cell(ws, 2, j+2, str(yr))

    # Calcular margen bruto y utilidades
    resultados = []
    for j in range(5):
        ventas = ing[j]['total']
        c_prod = subtot_prod[j]
        c_comm = subtot_comm[j]
        costos_directos = c_prod + c_comm
        margen_bruto = ventas - costos_directos
        g_admin = egr[j]['nomina']
        u_oper = margen_bruto - g_admin
        # Imporrenta 35% si hay utilidad, ajuste por primeros años (exención Ley 1429)
        impuesto = int(u_oper * 0.20) if u_oper > 0 and j >= 2 else 0
        u_neta = u_oper - impuesto
        resultados.append({
            'ventas': ventas,
            'c_prod': c_prod,
            'c_comm': c_comm,
            'costos_directos': costos_directos,
            'margen_bruto': margen_bruto,
            'g_admin': g_admin,
            'u_oper': u_oper,
            'impuesto': impuesto,
            'u_neta': u_neta,
        })

    filas = [
        ("TOTAL VENTAS O INGRESOS",          "ventas",         VERDE_CLARO,   True),
        ("(-) Costos de producción y logística", "c_prod",      GRIS_CLARO,    False),
        ("(-) Costos comercialización y ventas", "c_comm",      BLANCO,        False),
        ("COSTOS DIRECTOS TOTALES",           "costos_directos",AMARILLO,      True),
        ("= MARGEN BRUTO DE VENTAS",          "margen_bruto",   VERDE_CLARO,   True),
        ("(-) Gastos de administración (Nómina)", "g_admin",    GRIS_CLARO,    False),
        ("= UTILIDAD OPERACIONAL (UAII)",     "u_oper",         AMARILLO,      True),
        ("(-) Impuesto de renta (0 % años 1–2; 20 % año 3+)", "impuesto", BLANCO, False),
        ("★  UTILIDAD NETA",                  "u_neta",         NARANJA_CLARO, True),
    ]

    for i, (label, key, bg, bold) in enumerate(filas):
        r = 3 + i
        fc = NARANJA if "★" in label else ("1B5E20" if "MARGEN" in label or "VENTAS" in label else "212121")
        neg = key in ("c_prod","c_comm","costos_directos","g_admin","impuesto")
        detail_cell(ws, r, 1, label, bg, bold=bold, font_color=fc)
        for j in range(5):
            v = resultados[j][key]
            if neg:
                v_show = -abs(v)   # mostrar negativo
            else:
                v_show = v
            fmt_str = PESO if v >= 0 else f'-{PESO}'
            # usar formato condicional simple
            c = ws.cell(row=r, column=j+2, value=v_show)
            c.fill = _fill(bg)
            c.font = Font(bold=bold, color=fc, size=10, name="Calibri")
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = borde()
            c.number_format = '#,##0;(#,##0)'

    # % Margen (sin merge para poder escribir todas las columnas)
    r_mg = 3 + len(filas)
    detail_cell(ws, r_mg, 1, "% Margen Bruto sobre Ventas", GRIS_CLARO, bold=False)
    for j, r_data in enumerate(resultados):
        mg = r_data['margen_bruto'] / r_data['ventas'] if r_data['ventas'] else 0
        detail_cell(ws, r_mg, j+2, mg, GRIS_CLARO, align="right", number_fmt=PCT)


# ═══════════════════════════════════════════════════════════════════════════
# MAIN — Ensamblar el workbook
# ═══════════════════════════════════════════════════════════════════════════
def main():
    # Crear workbook nuevo desde cero (evita problema con MergedCells del original)
    wb = openpyxl.Workbook()

    # Eliminar hoja default
    wb.remove(wb.active)

    ing      = proj_ingresos()
    egr_list = proj_egresos(ing)

    ws_pv  = wb.create_sheet("PROYECCION DE VENTAS A 5 AÑOS")
    ws_cf  = wb.create_sheet("COSTOS DE FABRICACION MENSUAL")
    ws_pp  = wb.create_sheet("PLANTA DE PERSONAL")
    ws_inv = wb.create_sheet("INVERSION TOTAL")
    ws_eg  = wb.create_sheet("FLUJO DE EGRESOS")
    ws_in  = wb.create_sheet("FLUJO DE INGRESOS")
    ws_er  = wb.create_sheet("ESTADO DE RESULTADOS")

    write_proyeccion_ventas(ws_pv, ing)
    write_costos_fabricacion(ws_cf)
    write_planta_personal(ws_pp)
    write_inversion_total(ws_inv)
    subtot_prod, subtot_comm = write_flujo_egresos(ws_eg, ing, egr_list)
    write_flujo_ingresos(ws_in, ing)
    write_estado_resultados(ws_er, ing, egr_list, subtot_prod, subtot_comm)

    out = 'PLAN_FINANCIERO_FLORACAMPO_2025.xlsx'
    wb.save(out)

    # ── Verificación ──────────────────────────────────────────────────────
    print("=" * 60)
    print("  FLORACAMPO — VERIFICACIÓN PLAN FINANCIERO")
    print("=" * 60)
    for j, r in enumerate(ing):
        print(f"  [{r['año']}] Ventas: ${r['total']:>18,.0f} COP")
    print()
    for j in range(5):
        nom = egr_list[j]['nomina']
        mat = egr_list[j]['mat_prima']
        comb = egr_list[j]['combustible']
        print(f"  [{YEARS[j]}] Nómina: ${nom:>15,.0f} | Mat: ${mat:>15,.0f} | Combustible: ${comb:>12,.0f}")
    print()
    print(f"  Inversión inicial 2025: ${sum([120_000_000,150_000_000,130_000_000,16_000_000,12_000_000,30_000_000,8_500_000,68_000_000,15_000_000]):,.0f} COP")
    print()
    print(f"  ✅ Archivo guardado: {out}")

if __name__ == "__main__":
    main()
